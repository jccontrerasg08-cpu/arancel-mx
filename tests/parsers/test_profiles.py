from openpyxl import Workbook
import pytest

from arancel_mx.parsers.profiles import resolve_workbook_profile
from arancel_mx.parsers.workbooks import probe_workbook


def make_workbook(tmp_path, name, rows):
    path = tmp_path / name
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def test_resolves_ligie_profile_from_registered_aliases(tmp_path):
    path = make_workbook(
        tmp_path,
        "ligie.xlsx",
        [
            ["nota"],
            ["Fracción", "Descripción", "Unidad", "IGI", "IGE"],
            ["01012101", "Reproductores de raza pura.", "Cbza", "10", "Ex."],
        ],
    )

    resolved = resolve_workbook_profile(probe_workbook(path), "ligie_snapshot")

    assert resolved.parser_version == "ligie-profile-1"
    assert resolved.profile.header_row == 2
    assert resolved.profile.columns["code"] == "Fracción"
    assert resolved.profile.columns["unit_name"] == "Unidad"


def test_ligie_profile_prefers_unique_more_specific_registered_candidate(tmp_path):
    path = tmp_path / "current-ligie.xlsx"
    workbook = Workbook()
    fa = workbook.active
    fa.title = "FA"
    for _ in range(6):
        fa.append([None])
    fa.append(["Fracción Arancelaria", "Descripción", "Unidad de Medida"])
    fa.append(["01012101", "Reproductores", "Cabeza"])
    nico = workbook.create_sheet("NICO")
    for _ in range(6):
        nico.append([None])
    nico.append(["FRACCIÓN ARANCELARIA", "DESCRIPCIÓN"])
    nico.append(["01012101", "Reproductores"])
    workbook.save(path)

    resolved = resolve_workbook_profile(probe_workbook(path), "ligie_snapshot")

    assert resolved.profile.sheet == "FA"
    assert resolved.profile.header_row == 7
    assert resolved.profile.columns["unit_name"] == "Unidad de Medida"


def test_resolves_nico_profile_from_registered_aliases(tmp_path):
    path = make_workbook(
        tmp_path,
        "nico.xlsx",
        [
            ["Fracción Arancelaria", "NICO", "Descripción NICO"],
            ["01012101", "00", "Reproductores"],
        ],
    )

    resolved = resolve_workbook_profile(probe_workbook(path), "nico_snapshot")

    assert resolved.profile.columns == {
        "fraccion8": "Fracción Arancelaria",
        "nico2": "NICO",
        "description": "Descripción NICO",
    }


def test_ambiguous_profile_fails_closed_with_candidate_locations(tmp_path):
    path = make_workbook(
        tmp_path,
        "ambiguous.xlsx",
        [
            ["Fracción", "Descripción", "IGI", "IGE"],
            ["01012101", "A", "10", "Ex."],
            ["Fracción", "Descripción", "IGI", "IGE"],
        ],
    )

    with pytest.raises(ValueError) as error:
        resolve_workbook_profile(probe_workbook(path), "ligie_snapshot")

    message = str(error.value)
    assert "ambiguous workbook profile: ligie_snapshot" in message
    assert "Datos!1" in message
    assert "Datos!3" in message


def test_unknown_profile_fails_closed(tmp_path):
    path = make_workbook(tmp_path, "unknown.xlsx", [["foo", "bar"], ["1", "2"]])

    with pytest.raises(ValueError, match="unknown workbook profile"):
        resolve_workbook_profile(probe_workbook(path), "ligie_snapshot")
