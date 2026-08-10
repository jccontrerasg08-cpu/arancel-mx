from pathlib import Path
import io
import json
import tempfile
import unittest
from unittest.mock import patch
from datetime import date
from decimal import Decimal

import duckdb
from openpyxl import Workbook

from src.comex.arancel_mx import (
    PUBLIC_COLUMNS,
    code_level,
    consolidate_records,
    derive_name,
    format_code,
    normalize_code,
    parse_duty,
    record_id,
    semantic_record_hash,
    stage_rows,
    validate_staging,
    promote_staging,
)
from src.comex.db import connect, init_db
from src.comex.etl import SniceNicoSource
from src.comex.arancel_sources import (
    _hierarchy_entries_from_table,
    _chapter_entries_from_text,
    _chapter_entries_from_pages,
    discover_official_documents,
    parse_ligie_workbook,
    parse_ligie_pdf_hierarchy,
    parse_nico_workbook,
)
from src.comex.arancel_build import export_arancel_release, materialize_arancel
from src.comex.arancel_release import (
    _apply_observed_updates,
    _cached_hierarchy,
    build_arancel_release,
)
from src.comex.diputados_ligie import parse_ligie_ledger
from src.comex.arancel_reconcile import reconcile_legal_instruments
from comex import main


class PromotionTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.path = Path(self.temp.name) / "promotion.duckdb"
        init_db(self.path)
        self.conn = duckdb.connect(str(self.path))

    def tearDown(self):
        self.conn.close()
        self.temp.cleanup()

    def test_ambiguous_legal_row_is_quarantined_and_blocks_promotion(self):
        stage_rows(self.conn, [{
            "capture_id": "cap-1", "dataset_key": "ligie", "document_role": "ligie_current",
            "sheet_name": "Datos", "source_row_number": 2, "parser_version": "1",
            "raw": {"code": "0101210?"}, "normalized": {"code": "0101210?"},
        }])
        report = validate_staging(self.conn)
        self.assertFalse(report.publishable)
        self.assertEqual(report.quarantined[0].reason_code, "ambiguous_code")
        with self.assertRaisesRegex(ValueError, "quarantine"):
            promote_staging(self.conn)

    def test_valid_fraction_promotes_to_typed_table(self):
        stage_rows(self.conn, [{
            "capture_id": "cap-2", "dataset_key": "ligie", "document_role": "ligie_current",
            "sheet_name": "Datos", "source_row_number": 3, "parser_version": "1",
            "raw": {"code": "01012101"},
            "normalized": {
                "code": "01012101", "description": "Reproductores de raza pura",
                "ligie_version": "2026-04-23", "source_document_id": "doc-1",
                "validity_basis": "legal",
            },
        }])
        self.assertTrue(validate_staging(self.conn).publishable)
        summary = promote_staging(self.conn)
        self.assertEqual(summary.tariff_fractions, 1)
        self.assertEqual(self.conn.execute("SELECT code FROM tariff_fraction").fetchone()[0], "01012101")


class ReconciliationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        fixture = Path("tests/fixtures/diputados/ligie_2022.html").read_text(encoding="utf-8")
        cls.ledger = parse_ligie_ledger(
            fixture, "https://www.diputados.gob.mx/LeyesBiblio/ref/ligie_2022.htm"
        )
        cls.dof = (
            {"document_id": "dof-law", "role": "law_reform", "published_at": "2025-12-29"},
            {"document_id": "dof-tariff", "role": "tariff_decree", "published_at": "2026-04-23"},
        )
        cls.snice = (
            {"document_id": "nico-current", "role": "nico_agreement"},
            {"document_id": "proposal-1", "role": "nico_proposal"},
            {"document_id": "indicator-1", "role": "weighted_tariff_indicator"},
        )

    def test_diputados_date_requires_matching_dof_document(self):
        report = reconcile_legal_instruments(self.ledger, (), self.snice)
        self.assertFalse(report.publishable)
        self.assertIn("missing_dof_evidence", report.error_codes)

    def test_proposals_and_indicators_never_enter_legal_tables(self):
        result = reconcile_legal_instruments(self.ledger, self.dof, self.snice)
        self.assertTrue(result.publishable)
        self.assertFalse(set(result.legal_document_ids) & set(result.proposal_document_ids))
        self.assertFalse(set(result.legal_document_ids) & set(result.indicator_document_ids))


class ArancelSchemaTests(unittest.TestCase):
    def test_init_db_creates_normalized_tables_and_preserves_catalog_projection(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "arancel.duckdb"
            init_db(path)
            init_db(path)
            with connect(path) as conn:
                tables = {row[0] for row in conn.execute("SHOW TABLES").fetchall()}
                columns = [
                    row[1]
                    for row in conn.execute("PRAGMA table_info('tariff_fraction')").fetchall()
                ]

        self.assertTrue(
            {
                "source_document",
                "hs_code",
                "tariff_fraction",
                "nico",
                "tariff_rate",
                "canonical_record",
                "record_provenance",
                "dataset_release",
                "catalog_tariff_fraction",
                "catalog_tariff_nico",
                "catalog_tariff_rate",
            }.issubset(tables)
        )
        self.assertIn("classification_effective_from", columns)
        self.assertIn("ligie_version", columns)

    def test_init_db_moves_existing_catalog_rows_before_installing_canonical_tables(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "legacy.duckdb"
            conn = duckdb.connect(str(path))
            conn.execute(
                """
                CREATE TABLE tariff_fraction (
                    code VARCHAR PRIMARY KEY, fraccion8 VARCHAR, nico10 VARCHAR,
                    hs2 VARCHAR, hs4 VARCHAR, hs6 VARCHAR, description VARCHAR,
                    source_code VARCHAR, country_scope VARCHAR, source_file VARCHAR,
                    source_url VARCHAR, loaded_at TIMESTAMP
                );
                INSERT INTO tariff_fraction VALUES
                    ('01012101', '01012101', NULL, '01', '0101', '010121',
                     'Caballos.', 'snice-nico', 'MX', 'old.xlsx',
                     'https://example.test/old.xlsx', TIMESTAMP '2021-02-24 00:00:00');

                CREATE TABLE tariff_nico (
                    nico10 VARCHAR PRIMARY KEY, fraccion8 VARCHAR, nico VARCHAR,
                    description VARCHAR, source_code VARCHAR, source_file VARCHAR,
                    loaded_at TIMESTAMP
                );
                INSERT INTO tariff_nico VALUES
                    ('0101210100', '01012101', '00', 'Caballos.',
                     'snice-nico', 'old.xlsx', TIMESTAMP '2021-02-24 00:00:00');

                CREATE TABLE tariff_rate (
                    rate_id VARCHAR PRIMARY KEY, code VARCHAR, tax_code VARCHAR,
                    tax_name VARCHAR, import_rate VARCHAR, export_rate VARCHAR,
                    unit_name VARCHAR, effective_from DATE, effective_to DATE,
                    source_code VARCHAR, source_file VARCHAR, loaded_at TIMESTAMP
                );
                INSERT INTO tariff_rate VALUES
                    ('old-rate', '01012101', 'general', 'General', '10', 'Ex.',
                     'Cbza', NULL, NULL, 'legacy', 'old.xlsx',
                     TIMESTAMP '2021-02-24 00:00:00');
                """
            )
            conn.close()

            init_db(path)

            with connect(path) as conn:
                counts = conn.execute(
                    """
                    SELECT
                        (SELECT COUNT(*) FROM catalog_tariff_fraction),
                        (SELECT COUNT(*) FROM catalog_tariff_nico),
                        (SELECT COUNT(*) FROM catalog_tariff_rate)
                    """
                ).fetchone()
                canonical_rate_columns = {
                    row[1]
                    for row in conn.execute("PRAGMA table_info('tariff_rate')").fetchall()
                }

        self.assertEqual(counts, (1, 1, 1))
        self.assertIn("rate_revision_id", canonical_rate_columns)


class CanonicalPrimitiveTests(unittest.TestCase):
    def test_codes_are_validated_and_formatted_without_silent_padding(self):
        self.assertEqual(normalize_code("0101.21.01"), "01012101")
        self.assertEqual(normalize_code(0, component_width=2), "00")
        self.assertEqual(code_level("0101210100"), "nico10")
        self.assertEqual(format_code("0101210100"), "0101.21.01 00")
        with self.assertRaises(ValueError):
            normalize_code("1012101")
        with self.assertRaises(ValueError):
            normalize_code(10121)

    def test_name_uses_first_boundary_and_unicode_word_limit(self):
        self.assertEqual(
            derive_name("Caballos reproductores; los demás."),
            "Caballos reproductores",
        )
        long_text = "Árbol " * 30
        derived = derive_name(long_text)
        self.assertLessEqual(len(derived), 120)
        self.assertFalse(derived.endswith(" "))
        self.assertTrue(derived.startswith("Árbol"))

    def test_duty_keeps_literal_and_does_not_invent_percentages(self):
        self.assertEqual(parse_duty("10%"), ("ad_valorem", Decimal("10"), "10%"))
        self.assertEqual(parse_duty("Ex."), ("exento", Decimal("0"), "Ex."))
        self.assertEqual(parse_duty("Prohibida"), ("prohibida", None, "Prohibida"))
        self.assertEqual(
            parse_duty("0.36 USD/Kg"),
            ("especifica", None, "0.36 USD/Kg"),
        )
        self.assertEqual(
            parse_duty("10% + 0.36 USD/Kg"),
            ("compuesta", None, "10% + 0.36 USD/Kg"),
        )
        self.assertEqual(
            parse_duty("AE (2 Dls EUA por Pza)"),
            ("especifica", None, "AE (2 Dls EUA por Pza)"),
        )
        self.assertEqual(parse_duty("según decreto"), ("desconocida", None, "según decreto"))
        self.assertEqual(parse_duty(None), (None, None, None))

    def test_semantic_hash_ignores_release_transport_metadata(self):
        row = {
            "level": "fraccion8",
            "code": "01012101",
            "description": "Caballos.",
            "name": "Caballos",
            "hs2": "01",
            "hs4": "0101",
            "hs6": "010121",
            "fraccion8": "01012101",
            "nico2": None,
            "nico10": None,
            "unit_code": "Cbza",
            "unit_name": "Cabeza",
            "igi_text": "10",
            "igi_kind": "ad_valorem",
            "igi_value": Decimal("10"),
            "ige_text": "Ex.",
            "ige_kind": "exento",
            "ige_value": Decimal("0"),
            "ligie_version": "LIGIE-2022",
            "validity_basis": "legal",
            "classification_effective_from": date(2022, 12, 12),
            "classification_effective_to": None,
            "rate_effective_from": date(2022, 12, 12),
            "rate_effective_to": None,
            "effective_from": date(2022, 12, 12),
            "effective_to": None,
            "dataset_version": "2026.08.09",
            "retrieved_at": "2026-08-09T12:00:00Z",
            "primary_source_url": "https://example.test/one",
        }
        changed_transport = {
            **row,
            "dataset_version": "2026.08.10",
            "retrieved_at": "2026-08-10T12:00:00Z",
            "primary_source_url": "https://example.test/two",
        }

        self.assertEqual(
            semantic_record_hash(row),
            semantic_record_hash(changed_transport),
        )

        identity = {
            **row,
            "record_version": 1,
            "effective_from": date(2022, 12, 12),
            "effective_to": None,
        }
        self.assertEqual(len(record_id(identity)), 64)


class ConsolidationTests(unittest.TestCase):
    def test_intersects_classification_and_parent_rate_intervals_deterministically(self):
        classifications = [
            {
                "classification_id": "hs6-a",
                "level": "hs6",
                "code": "010121",
                "description": "Caballos reproductores de raza pura.",
                "ligie_version": "LIGIE-2022",
                "validity_basis": "legal",
                "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None,
                "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7),
                "source_document_id": "doc-classification",
            },
            {
                "classification_id": "fraction-a",
                "level": "fraccion8",
                "code": "01012101",
                "description": "Caballos reproductores de raza pura.",
                "ligie_version": "LIGIE-2022",
                "validity_basis": "legal",
                "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None,
                "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7),
                "source_document_id": "doc-classification",
            },
            {
                "classification_id": "nico-a",
                "level": "nico10",
                "code": "0101210100",
                "description": "Reproductores de raza pura.",
                "ligie_version": "LIGIE-2022",
                "validity_basis": "legal",
                "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None,
                "published_at": date(2022, 7, 1),
                "updated_at": date(2022, 7, 1),
                "source_document_id": "doc-nico",
            },
        ]
        rates = [
            {
                "rate_revision_id": "rate-a",
                "code": "01012101",
                "unit_code": "Cbza",
                "unit_name": "Cabeza",
                "igi_text": "10",
                "igi_kind": "ad_valorem",
                "igi_value": Decimal("10"),
                "ige_text": "Ex.",
                "ige_kind": "exento",
                "ige_value": Decimal("0"),
                "ligie_version": "LIGIE-2022",
                "rate_effective_from": date(2022, 12, 12),
                "rate_effective_to": date(2023, 6, 30),
                "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7),
                "source_document_id": "doc-rate-base",
            },
            {
                "rate_revision_id": "rate-b",
                "code": "01012101",
                "unit_code": "Cbza",
                "unit_name": "Cabeza",
                "igi_text": "15",
                "igi_kind": "ad_valorem",
                "igi_value": Decimal("15"),
                "ige_text": "Ex.",
                "ige_kind": "exento",
                "ige_value": Decimal("0"),
                "ligie_version": "LIGIE-2022",
                "rate_effective_from": date(2023, 7, 1),
                "rate_effective_to": None,
                "published_at": date(2023, 6, 30),
                "updated_at": date(2023, 6, 30),
                "source_document_id": "doc-rate-change",
            },
        ]
        release = {
            "dataset_version": "2026.08.09",
            "schema_version": "1.0.0",
            "effective_as_of": date(2026, 8, 9),
        }

        first = consolidate_records(classifications, rates, release)
        second = consolidate_records(classifications, rates, release)

        self.assertEqual(first, second)
        hs_rows = [row for row in first if row["level"] == "hs6"]
        fraction_rows = [row for row in first if row["level"] == "fraccion8"]
        nico_rows = [row for row in first if row["level"] == "nico10"]
        self.assertEqual(len(hs_rows), 1)
        self.assertIsNone(hs_rows[0]["igi_value"])
        self.assertIsNone(hs_rows[0]["values_from_level"])
        self.assertEqual([row["igi_value"] for row in fraction_rows], [Decimal("10"), Decimal("15")])
        self.assertEqual([row["igi_value"] for row in nico_rows], [Decimal("10"), Decimal("15")])
        self.assertEqual([row["values_from_level"] for row in nico_rows], ["fraccion8", "fraccion8"])
        self.assertEqual([row["record_version"] for row in fraction_rows], [1, 2])
        self.assertEqual([row["is_current"] for row in fraction_rows], [False, True])
        self.assertEqual(
            nico_rows[0]["source_document_ids_json"],
            '["doc-classification","doc-nico","doc-rate-base"]',
        )
        self.assertEqual(nico_rows[0]["source_count"], 3)
        self.assertTrue(all(len(row["record_hash"]) == 64 for row in first))
        self.assertTrue(all(len(row["record_id"]) == 64 for row in first))

    def test_nico_requires_a_contemporaneous_parent_fraction(self):
        classification = {
            "level": "nico10",
            "code": "0101210100",
            "description": "Reproductores.",
            "ligie_version": "LIGIE-2022",
            "validity_basis": "legal",
            "classification_effective_from": date(2022, 12, 12),
            "classification_effective_to": None,
            "source_document_id": "doc-nico",
        }
        rate = {
            "code": "01012101",
            "igi_text": "10",
            "igi_value": Decimal("10"),
            "ligie_version": "LIGIE-2022",
            "rate_effective_from": date(2022, 12, 12),
            "rate_effective_to": None,
            "source_document_id": "doc-rate",
        }
        release = {
            "dataset_version": "2026.08.09",
            "schema_version": "1.0.0",
            "effective_as_of": date(2026, 8, 9),
        }

        self.assertEqual(consolidate_records([classification], [rate], release), [])

    def test_null_start_is_current_only_for_an_observed_snapshot(self):
        base = {
            "level": "hs2",
            "code": "01",
            "description": "Animales vivos.",
            "ligie_version": "LIGIE-2022",
            "classification_effective_from": None,
            "classification_effective_to": None,
            "source_document_id": "doc-base",
        }
        release = {
            "dataset_version": "2026.08.09",
            "schema_version": "1.0.0",
            "effective_as_of": date(2026, 8, 9),
        }

        legal = consolidate_records([{**base, "validity_basis": "legal"}], [], release)
        observed = consolidate_records(
            [{**base, "validity_basis": "observed_snapshot"}], [], release
        )

        self.assertFalse(legal[0]["is_current"])
        self.assertTrue(observed[0]["is_current"])


class CanonicalDatabaseTests(unittest.TestCase):
    def test_materializes_normalized_tables_and_exact_public_view_atomically(self):
        source_documents = [
            {
                "source_document_id": "doc-base",
                "authority": "Secretaría de Economía",
                "publication_venue": "SNICE",
                "title": "LIGIE",
                "source_url": "https://www.snice.gob.mx/ligie.xlsx",
                "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "sha256": "a" * 64,
                "published_at": date(2022, 6, 7),
                "effective_from": date(2022, 12, 12),
                "effective_to": None,
                "observed_at": date(2026, 8, 9),
                "retrieved_at": "2026-08-09T12:00:00",
            },
            {
                "source_document_id": "doc-nico",
                "authority": "Secretaría de Economía",
                "publication_venue": "SNICE",
                "title": "NICO",
                "source_url": "https://www.snice.gob.mx/nico.xlsx",
                "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "sha256": "b" * 64,
                "published_at": date(2022, 7, 1),
                "effective_from": date(2022, 12, 12),
                "effective_to": None,
                "observed_at": date(2026, 8, 9),
                "retrieved_at": "2026-08-09T12:01:00",
            },
        ]
        classifications = [
            {
                "classification_id": "hs2-a", "level": "hs2", "code": "01",
                "description": "Animales vivos.", "ligie_version": "LIGIE-2022",
                "validity_basis": "legal", "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None, "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7), "source_document_id": "doc-base",
            },
            {
                "classification_id": "hs4-a", "level": "hs4", "code": "0101",
                "description": "Caballos, asnos, mulos y burdéganos, vivos.",
                "ligie_version": "LIGIE-2022", "validity_basis": "legal",
                "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None, "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7), "source_document_id": "doc-base",
            },
            {
                "classification_id": "hs6-a", "level": "hs6", "code": "010121",
                "description": "Reproductores de raza pura.", "ligie_version": "LIGIE-2022",
                "validity_basis": "legal", "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None, "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7), "source_document_id": "doc-base",
            },
            {
                "classification_id": "fraction-a",
                "level": "fraccion8",
                "code": "01012101",
                "description": "Caballos reproductores de raza pura.",
                "ligie_version": "LIGIE-2022",
                "validity_basis": "legal",
                "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None,
                "published_at": date(2022, 6, 7),
                "updated_at": date(2022, 6, 7),
                "source_document_id": "doc-base",
            },
            {
                "classification_id": "nico-a",
                "level": "nico10",
                "code": "0101210100",
                "description": "Reproductores de raza pura.",
                "ligie_version": "LIGIE-2022",
                "validity_basis": "legal",
                "classification_effective_from": date(2022, 12, 12),
                "classification_effective_to": None,
                "published_at": date(2022, 7, 1),
                "updated_at": date(2022, 7, 1),
                "source_document_id": "doc-nico",
            },
        ]
        rates = [{
            "rate_revision_id": "rate-a",
            "code": "01012101",
            "unit_code": "Cbza",
            "unit_name": "Cabeza",
            "igi_text": "10",
            "igi_kind": "ad_valorem",
            "igi_value": Decimal("10"),
            "ige_text": "Ex.",
            "ige_kind": "exento",
            "ige_value": Decimal("0"),
            "ligie_version": "LIGIE-2022",
            "rate_effective_from": date(2022, 12, 12),
            "rate_effective_to": None,
            "published_at": date(2022, 6, 7),
            "updated_at": date(2022, 6, 7),
            "source_document_id": "doc-base",
        }]
        release = {
            "dataset_version": "2026.08.09",
            "schema_version": "1.0.0",
            "ligie_version": "LIGIE-2022",
            "effective_as_of": date(2026, 8, 9),
            "generated_at": "2026-08-09T12:05:00",
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "candidate.duckdb"
            init_db(path)
            with connect(path) as conn:
                result = materialize_arancel(
                    conn, source_documents, classifications, rates, release
                )
                columns = [row[0] for row in conn.execute("DESCRIBE arancel_mx").fetchall()]
                counts = conn.execute(
                    """
                    SELECT
                      (SELECT COUNT(*) FROM tariff_fraction),
                      (SELECT COUNT(*) FROM nico),
                      (SELECT COUNT(*) FROM tariff_rate),
                      (SELECT COUNT(*) FROM arancel_mx)
                    """
                ).fetchone()
                nico_row = conn.execute(
                    "SELECT code, igi_text, ige_text, primary_source_authority FROM arancel_mx WHERE level='nico10'"
                ).fetchone()

                with self.assertRaises(ValueError):
                    materialize_arancel(
                        conn,
                        source_documents,
                        [{**classifications[0], "source_document_id": "missing"}],
                        rates,
                        {**release, "dataset_version": "broken"},
                    )
                retained = conn.execute(
                    "SELECT dataset_version, COUNT(*) FROM arancel_mx GROUP BY dataset_version"
                ).fetchone()

        self.assertEqual(columns, list(PUBLIC_COLUMNS))
        self.assertEqual(counts, (1, 1, 1, 5))
        self.assertEqual(nico_row, ("0101210100", "10", "Ex.", "Secretaría de Economía"))
        self.assertEqual(result["row_count"], 5)
        self.assertEqual(result["validation_status"], "passed")
        self.assertEqual(retained, ("2026.08.09", 5))

    def test_exports_equivalent_deterministic_public_artifacts(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            database = root / "candidate.duckdb"
            output = root / "release"
            init_db(database)
            with connect(database) as conn:
                conn.execute(
                    """
                    CREATE OR REPLACE VIEW arancel_mx AS
                    SELECT 'id-1'::VARCHAR record_id, 1::INTEGER record_version,
                           true::BOOLEAN is_current, '01'::VARCHAR code,
                           '01'::VARCHAR formatted_code, 'hs2'::VARCHAR AS "level",
                           '01'::VARCHAR hs2, NULL::VARCHAR hs4, NULL::VARCHAR hs6,
                           NULL::VARCHAR fraccion8, NULL::VARCHAR nico2,
                           NULL::VARCHAR nico10, 'Animales vivos'::VARCHAR AS "name",
                           'Animales vivos.'::VARCHAR description,
                           true::BOOLEAN name_is_derived, NULL::VARCHAR unit_code,
                           NULL::VARCHAR unit_name, NULL::VARCHAR values_from_level,
                           NULL::VARCHAR igi_text, NULL::VARCHAR igi_kind,
                           10.500000::DECIMAL(18,6) igi_value, NULL::VARCHAR ige_text,
                           NULL::VARCHAR ige_kind, NULL::DECIMAL(18,6) ige_value,
                           'LIGIE-2022'::VARCHAR ligie_version,
                           '2026.08.09'::VARCHAR dataset_version,
                           '1.0.0'::VARCHAR schema_version, repeat('a',64)::VARCHAR record_hash,
                           'legal'::VARCHAR validity_basis,
                           DATE '2022-06-07' updated_at, DATE '2022-06-07' published_at,
                           DATE '2022-12-12' classification_effective_from,
                           NULL::DATE classification_effective_to,
                           NULL::DATE rate_effective_from, NULL::DATE rate_effective_to,
                           DATE '2022-12-12' effective_from, NULL::DATE effective_to,
                           DATE '2026-08-09' observed_at,
                           TIMESTAMP '2026-08-09 12:00:00' retrieved_at,
                           'doc-base'::VARCHAR primary_source_document_id,
                           'Secretaría de Economía'::VARCHAR primary_source_authority,
                           'https://www.snice.gob.mx/ligie.xlsx'::VARCHAR primary_source_url,
                           '["doc-base"]'::VARCHAR source_document_ids_json,
                           1::BIGINT source_count
                    """
                )
                conn.execute(
                    """
                    INSERT INTO dataset_release VALUES
                    ('2026.08.09', '1.0.0', 'LIGIE-2022', DATE '2026-08-09',
                     TIMESTAMP '2026-08-09 12:05:00', 1, 'passed',
                     '{"status":"passed"}', '["doc-base"]')
                    """
                )
                conn.execute(
                    """
                    INSERT INTO source_document VALUES
                    ('doc-base', 'Secretaría de Economía', 'SNICE', 'LIGIE',
                     'https://www.snice.gob.mx/ligie.xlsx', 'application/xlsx',
                     repeat('b',64), NULL, DATE '2022-06-07', DATE '2022-12-12',
                     NULL, DATE '2026-08-09', TIMESTAMP '2026-08-09 12:00:00')
                    """
                )
                conn.execute(
                    """
                    INSERT INTO canonical_record
                    SELECT * EXCLUDE (
                        primary_source_authority, primary_source_url,
                        source_document_ids_json, source_count
                    ) FROM arancel_mx
                    """
                )
                conn.execute(
                    "INSERT INTO record_provenance VALUES "
                    "('id-1', 'doc-base', 'base', true)"
                )

            manifest = export_arancel_release(database, output)
            json_bytes = (output / "arancel_mx.json").read_bytes()
            csv_bytes = (output / "arancel_mx.csv").read_bytes()
            payload = json.loads(json_bytes.decode("utf-8"))
            with duckdb.connect(str(output / "arancel_mx.duckdb"), read_only=True) as conn:
                copied = conn.execute("SELECT record_id, record_hash FROM arancel_mx").fetchall()
                public_tables = {
                    row[0]
                    for row in conn.execute(
                        "SELECT table_name FROM information_schema.tables "
                        "WHERE table_schema='main' AND table_type='BASE TABLE'"
                    ).fetchall()
                }
                public_views = {
                    row[0]
                    for row in conn.execute(
                        "SELECT table_name FROM information_schema.tables "
                        "WHERE table_schema='main' AND table_type='VIEW'"
                    ).fetchall()
                }

            self.assertEqual(
                {path.name for path in output.iterdir()},
                {"arancel_mx.duckdb", "arancel_mx.csv", "arancel_mx.json", "manifest.json", "SHA256SUMS"},
            )
            self.assertFalse(json_bytes.startswith(b"\xef\xbb\xbf"))
            self.assertNotIn(b"\n", csv_bytes.replace(b"\r\n", b""))
            self.assertEqual(list(payload[0]), list(PUBLIC_COLUMNS))
            self.assertEqual(payload[0]["retrieved_at"], "2026-08-09T12:00:00Z")
            self.assertEqual(payload[0]["igi_value"], 10.5)
            self.assertEqual(copied, [("id-1", "a" * 64)])
            self.assertEqual(public_views, {
                "arancel_mx", "arancel_mx_national_notes", "nico_proposals",
                "arancel_mx_weighted_indicators",
            })
            self.assertEqual(public_tables, {
                "source_document", "hs_code", "tariff_fraction", "nico",
                "tariff_rate", "canonical_record", "record_provenance",
                "dataset_release", "load_run",
                "nico_version", "nico_amendment", "nico_amendment_line",
                "nico_proposal_batch", "nico_proposal", "national_note",
                "national_note_version", "national_note_amendment",
                "national_note_applicability", "indicator_methodology",
                "weighted_tariff_indicator",
            })
            self.assertEqual(manifest["row_count"], 1)
            self.assertNotIn("local_path", manifest["source_documents"][0])
            self.assertEqual(
                manifest["source_documents"][0]["source_url"],
                "https://www.snice.gob.mx/ligie.xlsx",
            )
            self.assertEqual(set(manifest["artifact_sha256"]), {
                "arancel_mx.duckdb", "arancel_mx.csv", "arancel_mx.json"
            })


class ArancelCliTests(unittest.TestCase):
    def test_release_build_rejects_existing_destination_before_source_work(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output = root / "existing"
            output.mkdir()
            with patch(
                "src.comex.arancel_release._capture_current_sources"
            ) as capture:
                with self.assertRaises(FileExistsError):
                    build_arancel_release(
                        root / "sources", output, "2026.08.09", date(2026, 8, 9)
                    )
        capture.assert_not_called()

    def test_observed_date_fills_updated_at_without_fabricating_legal_dates(self):
        documents = [{
            "source_document_id": "doc-1",
            "observed_at": date(2026, 8, 9),
        }]
        rows = [{
            "source_document_id": "doc-1",
            "updated_at": None,
            "published_at": None,
            "classification_effective_from": None,
        }]

        _apply_observed_updates(documents, rows)

        self.assertEqual(rows[0]["updated_at"], date(2026, 8, 9))
        self.assertIsNone(rows[0]["published_at"])
        self.assertIsNone(rows[0]["classification_effective_from"])

    def test_hierarchy_cache_is_keyed_by_source_hash_and_parser_version(self):
        source = {
            "sha256": "a" * 64,
            "filename": "ligie.pdf",
            "source_document_id": "doc-pdf",
        }
        row = {"level": "hs2", "code": "01", "description": "Animales vivos"}
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "ligie.pdf").write_bytes(b"pdf fixture")
            with patch(
                "src.comex.arancel_release.parse_ligie_pdf_hierarchy",
                return_value=[row],
            ) as parser:
                first = _cached_hierarchy(root, source)
                second = _cached_hierarchy(root, source)

        self.assertEqual(first, [row])
        self.assertEqual(second, [row])
        parser.assert_called_once()

    def test_arancel_build_returns_release_summary(self):
        expected = {"row_count": 5, "validation_status": "passed"}
        with patch("comex.build_arancel_release", return_value=expected) as build:
            with patch("sys.stdout", new_callable=io.StringIO) as stdout:
                code = main([
                    "arancel-build",
                    "--source-dir", "fixtures",
                    "--output-dir", "release",
                    "--dataset-version", "2026.08.09",
                    "--effective-as-of", "2026-08-09",
                ])

        self.assertEqual(code, 0)
        self.assertEqual(json.loads(stdout.getvalue()), expected)
        build.assert_called_once()


class _FakeResponse:
    def __init__(self, text, url):
        self.text = text
        self.url = url

    def raise_for_status(self):
        return None


class _FakeClient:
    def __init__(self, pages):
        self.pages = pages

    def get(self, url, timeout):
        return _FakeResponse(self.pages[url], url)


class OfficialSourceParserTests(unittest.TestCase):
    def test_pdf_chapter_parser_stops_before_notes_and_ignores_inline_references(self):
        text = """
        Capítulo 78 (Sic DOF 07-06-2022)
        Plomo y sus manufacturas

        Nota de subpartida.
        En este Capítulo, se entiende por plomo refinado...
        """
        self.assertEqual(
            _chapter_entries_from_text(text),
            [("78", "Plomo y sus manufacturas")],
        )

    def test_pdf_chapter_parser_carries_heading_across_repeated_page_header(self):
        pages = [
            "contenido anterior\nCapítulo 56\n",
            "LEY DE LOS IMPUESTOS GENERALES DE IMPORTACIÓN Y DE EXPORTACIÓN\n"
            "CÁMARA DE DIPUTADOS DEL H. CONGRESO DE LA UNIÓN\n"
            "412 de 893\nGuata, fieltro y tela sin tejer\nNotas.\ntexto de notas",
        ]
        self.assertEqual(
            _chapter_entries_from_pages(pages),
            [("56", "Guata, fieltro y tela sin tejer")],
        )

    def test_pdf_table_parser_handles_fragmented_columns_and_multiline_hs_text(self):
        table = [
            ["", "0301.91", "", "--", "", "Truchas (Salmo trutta,", "", ""],
            [None, None, None, None, None, "Oncorhynchus mykiss).", None, None],
            ["0301.91.01", None, None, "", "", "Truchas.", "", "Cbza"],
            ["", "03.02", "", "", "", "Pescado fresco o refrigerado.", "", ""],
        ]

        self.assertEqual(_hierarchy_entries_from_table(table), [
            ("030191", "Truchas (Salmo trutta, Oncorhynchus mykiss)."),
            ("0302", "Pescado fresco o refrigerado."),
        ])

    def test_pdf_parser_extracts_official_hs2_hs4_hs6_hierarchy(self):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "ligie.pdf"
            story = [
                Table([["Capítulo 01"], ["Animales vivos"]]),
                Spacer(1, 10),
                Table(
                    [
                        ["CÓDIGO", "", "DESCRIPCIÓN", "UNIDAD", "IMP.", "EXP."],
                        ["01.01", "", "Caballos, asnos, mulos y burdéganos, vivos.", "", "", ""],
                        ["0101.21", "--", "Reproductores de raza pura.", "", "", ""],
                        ["0101.21.01", "", "Reproductores de raza pura.", "Cbza", "10", "Ex."],
                    ],
                    colWidths=[70, 20, 280, 50, 40, 40],
                    style=TableStyle([
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ]),
                ),
            ]
            SimpleDocTemplate(str(path), pagesize=letter).build(story)

            rows = parse_ligie_pdf_hierarchy(
                path, "doc-pdf", "LIGIE-2022", date(2025, 12, 29), None
            )

        self.assertEqual([row["level"] for row in rows], ["hs2", "hs4", "hs6"])
        self.assertEqual([row["code"] for row in rows], ["01", "0101", "010121"])
        self.assertEqual(rows[0]["description"], "Animales vivos")

    def test_ligie_parser_reads_split_arancel_import_export_header(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "split-header.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Fracción Arancelaria", "Descripción", "Unidad de Medida", "Arancel %", None])
            sheet.append([None, None, None, "IMP.", "EXP."])
            sheet.append(["0101.21.01", "Reproductores.", "Cbza", "10", "Ex."])
            workbook.save(path)

            _classifications, rates = parse_ligie_workbook(
                path, "doc", "LIGIE-2022", None, None
            )

        self.assertEqual(rates[0]["igi_text"], "10")
        self.assertEqual(rates[0]["ige_text"], "Ex.")

    def test_ligie_parser_collapses_exact_duplicates_but_rejects_conflicts(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "duplicates.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["CÓDIGO", "DESCRIPCIÓN", "Unidad", "IMP.", "EXP."])
            sheet.append(["0101.21.01", "Caballos.", "Cbza", "10", "Ex."])
            sheet.append(["0101.21.01", "Caballos.", "Cbza", "10", "Ex."])
            workbook.save(path)
            classifications, rates = parse_ligie_workbook(
                path, "doc", "LIGIE-2022", None, None
            )
            self.assertEqual((len(classifications), len(rates)), (1, 1))

            sheet.append(["0101.21.01", "Descripción incompatible.", "Cbza", "10", "Ex."])
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "Conflicting classification"):
                parse_ligie_workbook(path, "doc", "LIGIE-2022", None, None)

    def test_ligie_parser_does_not_treat_nico_sheet_as_fraction_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mixed.xlsx"
            workbook = Workbook()
            fraction_sheet = workbook.active
            fraction_sheet.title = "FA"
            fraction_sheet.append(["CÓDIGO", "DESCRIPCIÓN", "Unidad", "IMP.", "EXP."])
            fraction_sheet.append(["0101.29.99", "Los demás.", "Pza", "20", "Ex."])
            nico_sheet = workbook.create_sheet("NICO")
            nico_sheet.append(["FRACCIÓN ARANCELARIA", "NICO", "DESCRIPCIÓN"])
            nico_sheet.append(["0101.29.99", "01", "Para saltos o carreras."])
            workbook.save(path)

            classifications, _rates = parse_ligie_workbook(
                path, "doc", "LIGIE-2022", None, None
            )

        self.assertEqual(len(classifications), 1)
        self.assertEqual(classifications[0]["description"], "Los demás.")

    def test_ligie_parser_preserves_hierarchy_description_unit_igi_and_ige(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "ligie.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["CÓDIGO", "DESCRIPCIÓN", "Unidad", "IMP.", "EXP."])
            sheet.append(
                [
                    "01.01",
                    "Caballos, asnos, mulos y burdéganos, vivos.",
                    None,
                    None,
                    None,
                ]
            )
            sheet.append(
                [
                    "0101.21.01",
                    "Caballos\nreproductores de raza pura.",
                    "Cbza",
                    "10",
                    "Ex.",
                ]
            )
            workbook.save(path)

            classifications, rates = parse_ligie_workbook(
                path,
                "doc-ligie",
                "LIGIE-2022",
                date(2022, 6, 7),
                date(2022, 12, 12),
            )

        self.assertEqual([row["code"] for row in classifications], ["0101", "01012101"])
        self.assertEqual(
            classifications[-1]["description"],
            "Caballos reproductores de raza pura.",
        )
        self.assertEqual(rates[0]["unit_code"], "Cbza")
        self.assertEqual(rates[0]["igi_text"], "10")
        self.assertEqual(rates[0]["igi_kind"], "ad_valorem")
        self.assertEqual(rates[0]["ige_kind"], "exento")

    def test_nico_parser_allows_only_component_padding(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "nico.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["FRACCIÓN ARANCELARIA", "NICO", "DESCRIPCIÓN"])
            sheet.append(["0101.21.01", 0, "Reproductores de raza pura."])
            sheet.append(["no válido", "A", "Debe ignorarse"])
            workbook.save(path)

            rows = parse_nico_workbook(
                path,
                "doc-nico",
                "LIGIE-2022",
                date(2022, 7, 1),
                date(2022, 12, 12),
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["nico10"], "0101210100")
        self.assertEqual(rows[0]["fraccion8"], "01012101")
        self.assertEqual(rows[0]["nico2"], "00")

    def test_discovery_accepts_only_official_document_links(self):
        ligie_url = "https://www.snice.gob.mx/ligie"
        nico_url = "https://www.snice.gob.mx/nico"
        modification_url = "https://www.snice.gob.mx/modificaciones"
        pages = {
            ligie_url: """
                <a href='/~oracle/SNICE_DOCS/LIGIE2022.xlsx'>LIGIE 2022 Excel</a>
                <a href='https://evil.example/ligie.xlsx'>Copia</a>
            """,
            nico_url: """
                <a href='/~oracle/SNICE_DOCS/NICO2022.xlsx'>NICO 2022 Excel</a>
            """,
            modification_url: """
                <a href='https://www.dof.gob.mx/nota_detalle.php?codigo=123'>DOF modificación</a>
                <a href='/~oracle/SNICE_DOCS/MODLIGIE2026.xlsx'>Cambios Excel 23 abril 2026</a>
            """,
        }

        documents = discover_official_documents(
            _FakeClient(pages),
            ligie_url,
            nico_url,
            modification_url,
            timeout_s=10,
        )

        self.assertEqual(
            [document["source_url"] for document in documents],
            [
                "https://www.snice.gob.mx/~oracle/SNICE_DOCS/LIGIE2022.xlsx",
                "https://www.snice.gob.mx/~oracle/SNICE_DOCS/NICO2022.xlsx",
                "https://www.dof.gob.mx/nota_detalle.php?codigo=123",
                "https://www.snice.gob.mx/~oracle/SNICE_DOCS/MODLIGIE2026.xlsx",
            ],
        )
        self.assertEqual(
            [document["kind"] for document in documents],
            ["ligie", "nico", "modification", "modification"],
        )

    def test_snice_source_turns_discovered_documents_into_download_tasks(self):
        ligie_url = "https://www.snice.gob.mx/ligie"
        nico_url = "https://www.snice.gob.mx/nico"
        modification_url = "https://www.snice.gob.mx/modificaciones"
        pages = {
            ligie_url: "<a href='/docs/LIGIE2022.xlsx'>LIGIE 2022 Excel</a>",
            nico_url: "<a href='/docs/NICO2022.xlsx'>NICO 2022 Excel</a>",
            modification_url: "<a href='/docs/MODLIGIE2026.xlsx'>Modificación 2026 Excel</a>",
        }
        source = SniceNicoSource(ligie_url, nico_url, modification_url)

        tasks = source.discover_tasks(_FakeClient(pages), timeout_s=10)

        self.assertEqual([task.relative_path for task in tasks], [
            "ligie-001.xlsx",
            "nico-001.xlsx",
            "modification-001.xlsx",
        ])
        self.assertEqual([task.extra["kind"] for task in tasks], [
            "ligie",
            "nico",
            "modification",
        ])

    def test_discovery_follows_year_pages_and_classifies_mixed_landing_page_files(self):
        ligie_url = "https://www.snice.gob.mx/ligie"
        nico_url = "https://www.snice.gob.mx/nico-modificaciones"
        modification_url = "https://www.snice.gob.mx/modificaciones"
        nico_year = "https://www.snice.gob.mx/ligie.nico22.mod24.html"
        modification_year = "https://www.snice.gob.mx/ligie.info22.mod26.html"
        obsolete_nico_year = "https://www.snice.gob.mx/ligie.nico22.mod21.html"
        obsolete_modification_year = "https://www.snice.gob.mx/ligie.info22.mod22.html"
        pages = {
            ligie_url: """
                <a href='/docs/FRACCIONESARANCELARIAS-LIGIE_20260420.xlsx'>fracciones arancelarias</a>
                <a href='/docs/NICO-MARZO24-LIGIE_20240404.xlsx'>NICO</a>
                <a href='/docs/ARANCEL-CUPO-2024-LIGIE.xlsx'>aranceles-cupos</a>
            """,
            nico_url: (
                f"<a href='{obsolete_nico_year}'></a>"
                f"<a href='{nico_year}'></a>"
            ),
            nico_year: "<a href='/docs/MODIFICACIONES-NICO-2024.xlsx'>Cambios NICO</a>",
            modification_url: (
                f"<a href='{obsolete_modification_year}'></a>"
                f"<a href='{modification_year}'></a>"
            ),
            modification_year: """
                <a href='/docs/MODIFICACIONES-ABRIL2026.xlsx'>Cambios Excel</a>
                <a href='/docs/CANASTABASICA-7MAENMIENDA.xlsx'>Canasta básica</a>
                <a href='https://www.dof.gob.mx/nota_detalle.php?codigo=456'>DOF 23 abril 2026</a>
            """,
        }

        documents = discover_official_documents(
            _FakeClient(pages), ligie_url, nico_url, modification_url, timeout_s=10
        )

        self.assertEqual(
            [(document["kind"], Path(document["source_url"].split("?")[0]).name) for document in documents],
            [
                ("ligie", "FRACCIONESARANCELARIAS-LIGIE_20260420.xlsx"),
                ("nico", "NICO-MARZO24-LIGIE_20240404.xlsx"),
                ("nico", "MODIFICACIONES-NICO-2024.xlsx"),
                ("modification", "MODIFICACIONES-ABRIL2026.xlsx"),
                ("modification", "CANASTABASICA-7MAENMIENDA.xlsx"),
                ("modification", "nota_detalle.php"),
            ],
        )


if __name__ == "__main__":
    unittest.main()
