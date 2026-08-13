from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator, Mapping


@dataclass(frozen=True)
class WorkbookProfile:
    sheet: str
    header_row: int
    columns: Mapping[str, str]
    forward_fill: tuple[str, ...] = ()
    section_column: str | None = None
    allowed_sections: tuple[str, ...] = ()
    data_row: int | None = None
    column_indices: Mapping[str, int] = field(default_factory=dict)


@dataclass(frozen=True)
class WorkbookProbe:
    sheet_names: tuple[str, ...]
    samples: Mapping[str, tuple[tuple[Any, ...], ...]]


@dataclass(frozen=True)
class StagingRow:
    domain: str
    source_document_id: str
    sheet: str
    row_number: int
    raw: Mapping[str, Any]
    normalized: Mapping[str, Any]


def _excel_engine(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return "xlrd"
    if suffix == ".xlsx":
        return "openpyxl"
    raise ValueError(f"unsupported workbook format: {path.suffix}")


def _cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _iter_openpyxl_rows(path: Path, sheet_name: str | None = None) -> Iterator[tuple[str, list[str]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        names = (sheet_name,) if sheet_name is not None else tuple(workbook.sheetnames)
        for name in names:
            for row in workbook[name].iter_rows(values_only=True):
                yield name, [_cell(item) for item in row]
    finally:
        workbook.close()


def _iter_xlrd_rows(path: Path, sheet_name: str | None = None) -> Iterator[tuple[str, list[str]]]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    names = (sheet_name,) if sheet_name is not None else tuple(book.sheet_names())
    for name in names:
        sheet = book.sheet_by_name(name)
        for row_index in range(sheet.nrows):
            yield name, [_cell(sheet.cell_value(row_index, column)) for column in range(sheet.ncols)]


def _iter_rows(path: Path, sheet_name: str | None = None) -> Iterator[tuple[str, list[str]]]:
    if _excel_engine(path) == "openpyxl":
        yield from _iter_openpyxl_rows(path, sheet_name)
        return
    yield from _iter_xlrd_rows(path, sheet_name)


def probe_workbook(path: Path, sample_rows: int = 20, sample_columns: int = 30) -> WorkbookProbe:
    if sample_rows < 1 or sample_columns < 1:
        raise ValueError("workbook probe bounds must be positive")
    samples: dict[str, list[tuple[Any, ...]]] = {}
    order: list[str] = []
    counts: dict[str, int] = {}
    for name, row in _iter_rows(path):
        if name not in samples:
            order.append(name)
            samples[name] = []
            counts[name] = 0
        if counts[name] >= sample_rows:
            continue
        samples[name].append(tuple(row[:sample_columns]))
        counts[name] += 1
    return WorkbookProbe(tuple(order), {name: tuple(rows) for name, rows in samples.items()})


def _apply_forward_fill(rows: list[dict[str, str]], profile: WorkbookProfile) -> list[dict[str, str]]:
    for logical in profile.forward_fill:
        if logical not in profile.columns:
            raise ValueError(f"forward_fill column is not registered: {logical}")
        last = ""
        for raw in rows:
            value = raw[logical]
            if str(value).strip():
                last = value
            elif last:
                raw[logical] = last
    return rows


def _read_profile(path: Path, profile: WorkbookProfile) -> list[dict[str, str]]:
    if profile.header_row < 1:
        raise ValueError("header_row is one-based")
    matrix = [row for _name, row in _iter_rows(path, profile.sheet)]

    if profile.column_indices:
        if set(profile.column_indices) != set(profile.columns):
            raise ValueError("column_indices must cover every registered logical column")
        if any(index < 0 for index in profile.column_indices.values()):
            raise ValueError("column_indices must be zero-based non-negative positions")
        data_row = profile.data_row or profile.header_row + 1
        if data_row <= profile.header_row:
            raise ValueError("data_row must follow header_row")
        width = max((len(row) for row in matrix), default=0)
        if max(profile.column_indices.values()) >= width:
            raise ValueError("registered workbook column position is out of bounds")
        start = data_row - 1
        rows = []
        for row in matrix[start:]:
            padded = row + [""] * (width - len(row))
            rows.append({logical: padded[index] for logical, index in profile.column_indices.items()})
        return _apply_forward_fill(rows, profile)

    headers = list(dict.fromkeys(profile.columns.values()))
    header_idx = profile.header_row - 1
    if header_idx >= len(matrix):
        raise ValueError(f"missing registered workbook columns: {sorted(headers)}")
    header_to_index: dict[str, int] = {}
    for index, cell in enumerate(matrix[header_idx]):
        if cell and cell not in header_to_index:
            header_to_index[cell] = index
    missing = set(headers).difference(header_to_index)
    if missing:
        raise ValueError(f"missing registered workbook columns: {sorted(missing)}")
    rows = []
    for row in matrix[header_idx + 1 :]:
        raw = {}
        for logical, header in profile.columns.items():
            index = header_to_index[header]
            raw[logical] = row[index] if index < len(row) else ""
        rows.append(raw)
    return _apply_forward_fill(rows, profile)


def _rows(path: Path, source: Mapping[str, Any], profile: WorkbookProfile):
    source_document_id = str(source.get("source_document_id", ""))
    if not source_document_id:
        raise ValueError("source_document_id is required")
    frame = _read_profile(path, profile)
    first_data_row = profile.data_row or profile.header_row + 1
    for offset, raw in enumerate(frame):
        if not any(str(value).strip() for value in raw.values()):
            continue
        if profile.section_column:
            section_value = str(raw.get(profile.section_column, "")).strip()
            if profile.allowed_sections and section_value not in profile.allowed_sections:
                continue
        yield source_document_id, first_data_row + offset, raw


def parse_ligie_workbook(
    path: Path, source: Mapping[str, Any], profile: WorkbookProfile
) -> list[StagingRow]:
    from arancel_mx.domain.normalization import normalize_code, parse_duty

    result = []
    for source_id, row_number, raw in _rows(path, source, profile):
        try:
            code = normalize_code(raw.get("code"))
        except ValueError as exc:
            raise ValueError("complete LIGIE code has invalid width") from exc
        if len(code) != 8:
            raise ValueError(f"complete LIGIE code has width {len(code)}; expected 8")
        igi = parse_duty(raw.get("igi")) if "igi" in raw else (None, None, None)
        ige = parse_duty(raw.get("ige")) if "ige" in raw else (None, None, None)
        normalized = {
            "code": code,
            "description": str(raw.get("description", "")).strip(),
            "igi_kind": igi[0], "igi_value": igi[1], "igi_text": igi[2],
            "ige_kind": ige[0], "ige_value": ige[1], "ige_text": ige[2],
        }
        if "unit_code" in raw:
            normalized["unit_code"] = str(raw["unit_code"]).strip() or None
        if "unit_name" in raw:
            normalized["unit_name"] = str(raw["unit_name"]).strip() or None
        result.append(StagingRow("legal", source_id, profile.sheet, row_number, raw, normalized))
    return result


def parse_nico_workbook(
    path: Path, source: Mapping[str, Any], profile: WorkbookProfile
) -> list[StagingRow]:
    from arancel_mx.domain.normalization import normalize_code

    result = []
    for source_id, row_number, raw in _rows(path, source, profile):
        if "nico10" in raw:
            try:
                nico10 = normalize_code(raw["nico10"])
            except ValueError as exc:
                raise ValueError("complete NICO code has invalid width") from exc
            if len(nico10) != 10:
                raise ValueError(f"complete NICO code has width {len(nico10)}; expected 10")
            fraccion8, nico2 = nico10[:8], nico10[8:]
        else:
            try:
                fraccion8 = normalize_code(raw.get("fraccion8"))
            except ValueError as exc:
                raise ValueError("complete LIGIE code has invalid width") from exc
            if len(fraccion8) != 8:
                raise ValueError(f"complete LIGIE code has width {len(fraccion8)}; expected 8")
            nico2 = normalize_code(raw.get("nico2"), component_width=2)
            nico10 = fraccion8 + nico2
        normalized = {
            "nico10": nico10,
            "fraccion8": fraccion8,
            "nico2": nico2,
            "description": str(raw.get("description", "")).strip(),
        }
        result.append(StagingRow("legal", source_id, profile.sheet, row_number, raw, normalized))
    return result


def parse_indicator_workbook(
    path: Path, source: Mapping[str, Any], profile: WorkbookProfile
) -> list[StagingRow]:
    result = []
    for source_id, row_number, raw in _rows(path, source, profile):
        normalized = {key: str(value).strip() for key, value in raw.items()}
        result.append(StagingRow("analytics", source_id, profile.sheet, row_number, raw, normalized))
    return result
