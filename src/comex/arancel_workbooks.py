from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook

from .arancel_mx import normalize_code, parse_duty


@dataclass(frozen=True)
class WorkbookProfile:
    sheet: str
    header_row: int
    columns: Mapping[str, str]
    forward_fill: tuple[str, ...] = ()
    section_column: str | None = None
    allowed_sections: tuple[str, ...] = ()


@dataclass(frozen=True)
class WorkbookProbe:
    sheet_names: tuple[str, ...]
    samples: Mapping[str, tuple[tuple[Any, ...], ...]]


@dataclass(frozen=True)
class StagingRow:
    domain: str
    source_document_id: str
    sheet: str
    row_number: int
    raw: Mapping[str, Any]
    normalized: Mapping[str, Any]


def probe_workbook(path: Path, sample_rows: int = 20, sample_columns: int = 30) -> WorkbookProbe:
    if sample_rows < 1 or sample_columns < 1:
        raise ValueError("workbook probe bounds must be positive")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        samples = {
            sheet.title: tuple(
                tuple(row[:sample_columns])
                for row in sheet.iter_rows(max_row=sample_rows, values_only=True)
            )
            for sheet in workbook.worksheets
        }
        return WorkbookProbe(tuple(workbook.sheetnames), samples)
    finally:
        workbook.close()


def _read_profile(path: Path, profile: WorkbookProfile) -> pd.DataFrame:
    if profile.header_row < 1:
        raise ValueError("header_row is one-based")
    headers = list(dict.fromkeys(profile.columns.values()))
    frame = pd.read_excel(
        path,
        sheet_name=profile.sheet,
        header=profile.header_row - 1,
        usecols=headers,
        dtype=str,
        keep_default_na=False,
        engine="openpyxl",
    )
    missing = set(headers).difference(frame.columns)
    if missing:
        raise ValueError(f"missing registered workbook columns: {sorted(missing)}")
    return frame


def _rows(path: Path, source: Mapping[str, Any], profile: WorkbookProfile):
    source_document_id = str(source.get("source_document_id", ""))
    if not source_document_id:
        raise ValueError("source_document_id is required")
    frame = _read_profile(path, profile)
    for offset, values in enumerate(frame.to_dict(orient="records"), start=1):
        raw = {logical: values[header] for logical, header in profile.columns.items()}
        if not any(str(value).strip() for value in raw.values()):
            continue
        yield source_document_id, profile.header_row + offset, raw


def parse_ligie_workbook(
    path: Path, source: Mapping[str, Any], profile: WorkbookProfile
) -> list[StagingRow]:
    result = []
    for source_id, row_number, raw in _rows(path, source, profile):
        try:
            code = normalize_code(raw.get("code"))
        except ValueError as exc:
            raise ValueError("complete LIGIE code has invalid width") from exc
        if len(code) != 8:
            raise ValueError(f"complete LIGIE code has width {len(code)}; expected 8")
        igi = parse_duty(raw.get("igi")) if "igi" in raw else (None, None, None)
        ige = parse_duty(raw.get("ige")) if "ige" in raw else (None, None, None)
        normalized = {
            "code": code,
            "description": str(raw.get("description", "")).strip(),
            "igi_kind": igi[0], "igi_value": igi[1], "igi_text": igi[2],
            "ige_kind": ige[0], "ige_value": ige[1], "ige_text": ige[2],
        }
        result.append(StagingRow("legal", source_id, profile.sheet, row_number, raw, normalized))
    return result


def parse_nico_workbook(
    path: Path, source: Mapping[str, Any], profile: WorkbookProfile
) -> list[StagingRow]:
    result = []
    for source_id, row_number, raw in _rows(path, source, profile):
        if "nico10" in raw:
            try:
                nico10 = normalize_code(raw["nico10"])
            except ValueError as exc:
                raise ValueError("complete NICO code has invalid width") from exc
            if len(nico10) != 10:
                raise ValueError(f"complete NICO code has width {len(nico10)}; expected 10")
            fraccion8, nico2 = nico10[:8], nico10[8:]
        else:
            try:
                fraccion8 = normalize_code(raw.get("fraccion8"))
            except ValueError as exc:
                raise ValueError("complete LIGIE code has invalid width") from exc
            if len(fraccion8) != 8:
                raise ValueError(f"complete LIGIE code has width {len(fraccion8)}; expected 8")
            nico2 = normalize_code(raw.get("nico2"), component_width=2)
            nico10 = fraccion8 + nico2
        normalized = {
            "nico10": nico10,
            "fraccion8": fraccion8,
            "nico2": nico2,
            "description": str(raw.get("description", "")).strip(),
        }
        result.append(StagingRow("legal", source_id, profile.sheet, row_number, raw, normalized))
    return result


def parse_indicator_workbook(
    path: Path, source: Mapping[str, Any], profile: WorkbookProfile
) -> list[StagingRow]:
    result = []
    for source_id, row_number, raw in _rows(path, source, profile):
        normalized = {key: str(value).strip() for key, value in raw.items()}
        result.append(StagingRow("analytics", source_id, profile.sheet, row_number, raw, normalized))
    return result
